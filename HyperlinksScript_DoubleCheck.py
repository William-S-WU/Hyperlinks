#William Schmitt
#requires openyxl and wget
import openpyxl
import os
import subprocess
import re

directory = os.getcwd()  # Working directory
print("Working Directory:", directory)

# Define file formats and corresponding paths for Google Sheets and Google Docs
file_formats = [
    ('xlsx', 'spreadsheets'),
    ('docx', 'document'),
    ('pdf', None)  # None here indicates using the original link for PDFs
]

# Function to transform Google Drive link into a valid export URL
def transform_google_drive_link(link, desired_format, doc_type):
    if doc_type is None:
        return link  # Use the original link for PDFs
    match = re.search(r'/d/([a-zA-Z0-9_-]+)/', link)
    if match:
        file_id = match.group(1)
        return f"https://docs.google.com/{doc_type}/d/{file_id}/export?format={desired_format}"
    return None

# Function to extract display text and transform hyperlinks from an Excel file
def extract_and_transform_hyperlinks(filename):
    wb = openpyxl.load_workbook(filename)
    hyperlinks = {}

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        sheet_hyperlinks = []

        for row in sheet.iter_rows():
            for cell in row:
                if cell.hyperlink:
                    if not cell.hyperlink.target.lower().startswith("mailto:"):
                        original_link = cell.hyperlink.target
                        display_text = cell.value or 'Unnamed'  # Ensure there is a display value
                        sanitized_filename = re.sub(r'\W+', '_', display_text)
                        sheet_hyperlinks.append((sanitized_filename, original_link))

        hyperlinks[sheet_name] = sheet_hyperlinks

    return hyperlinks

# Function to save hyperlinks to a text file
def save_hyperlinks_to_file(folder_path, sheet_name, hyperlinks):
    file_path = os.path.join(folder_path, f"{sheet_name}_download_links.txt")
    print(f"Saving download links to {file_path}")
    with open(file_path, "w", encoding="utf-8") as f:
        for filename, original_link in hyperlinks:
            f.write(f"{filename}\t{original_link}\n")

# Function to download files using wget with multiple formats
def download_from_text_file(text_file_path):
    with open(text_file_path, "r", encoding="utf-8") as f:
        links = f.readlines()

    for link in links:
        parts = link.strip().split('\t')
        if len(parts) == 2:
            original_filename, original_link = parts
            success = False

            for file_format, doc_type in file_formats:
                transformed_link = transform_google_drive_link(original_link, file_format, doc_type)
                if transformed_link:
                    destination_path = os.path.join(os.path.dirname(text_file_path), f"{original_filename}.{file_format}")
                    command = f'wget --no-check-certificate -O "{destination_path}" "{transformed_link}"'

                    print(f"Executing command: {command}")  # For debugging
                    result = subprocess.run(command, capture_output=True, text=True, shell=True)

                    if result.returncode == 0:
                        print(f"Downloaded {transformed_link} successfully as {destination_path}")
                        success = True
                        break
                    else:
                        print(f"Error downloading {transformed_link}: {result.stderr}")

                if success:
                    break

# Function to remove files with 0 bytes and empty subdirectories
def remove_zero_byte_files_and_empty_dirs(root_folder):
    print("Starting cleanup of zero-byte files and empty subdirectories.")  # Debug info

    for subdir, dirs, files in os.walk(root_folder):
        for file in files:
            file_path = os.path.join(subdir, file)
            if os.path.getsize(file_path) == 0:
                print(f"Deleting 0-byte file: {file_path}")
                os.remove(file_path)
        
        # Remove empty directories
        if not os.listdir(subdir):
            print(f"Removing empty directory: {subdir}")
            os.rmdir(subdir)
    print("Cleanup of zero-byte files and empty subdirectories completed.")

# Function to process each Excel file
def process_excel_files(directory):
    created_folder_paths = []
    for filename in os.listdir(directory):
        if filename.lower().endswith(".xlsx") or filename.lower().endswith(".xls"):
            print(f"Processing: {filename}")
            file_path = os.path.join(directory, filename)

            # Create a folder for the output text file and subfolders for each sheet
            base_name = os.path.splitext(filename)[0]
            folder_path = os.path.join(directory, base_name)
            try:
                os.makedirs(folder_path, exist_ok=True)
                print(f"Created folder: {folder_path}")
                created_folder_paths.append(folder_path)
            except OSError as e:
                print(f"Error creating folder {folder_path}: {e}")
                continue

            # Extract and transform hyperlinks
            hyperlinks = extract_and_transform_hyperlinks(file_path)
            print(f"Extracted and transformed hyperlinks from {filename}: {hyperlinks}")

            # Create subfolders for each sheet and save the hyperlinks
            for sheet_name, links in hyperlinks.items():
                sheet_folder_path = os.path.join(folder_path, sheet_name)
                print(f"Creating subfolder: {sheet_folder_path}")
                try:
                    os.makedirs(sheet_folder_path, exist_ok=True)
                    print(f"Created subfolder: {sheet_folder_path}")
                except OSError as e:
                    print(f"Error creating subfolder {sheet_folder_path}: {e}")
                    continue

                if links:
                    print(f"Hyperlinks for sheet {sheet_name}: {links}")
                    save_hyperlinks_to_file(sheet_folder_path, sheet_name, links)
                    download_from_text_file(os.path.join(sheet_folder_path, f"{sheet_name}_download_links.txt"))
    return created_folder_paths

# Function to find missing files
def find_missing_files(directory):
    print("Starting search for missing files.")

    # Find the input text file in the directory
    input_text_file = None
    for file in os.listdir(directory):
        if file.endswith(".txt"):
            if input_text_file is not None:
                raise ValueError("Multiple text files found. There should be only one text file.")
            input_text_file = file

    if input_text_file is None:
        raise ValueError("No text file found in the directory.")

    input_text_file_path = os.path.join(directory, input_text_file)
    output_log_file_path = os.path.join(directory, f"{os.path.splitext(input_text_file)[0]}_output.txt")

    # Read names from the text file (only the part before the tab)
    expected_files = []
    with open(input_text_file_path, "r", encoding="utf-8") as f:
        for line in f:
            name = line.split('\t')[0].strip()  # Extract the name part before the tab
            expected_files.append(name)

    print(expected_files)
    missing_files = []

    for i in range(len(expected_files)):
        k = 0
        for file in os.listdir(directory):
            print(file)
            file = os.path.splitext(file)[0]
            if expected_files[i] == file:
                k = 1
                break
        if k == 0:
            missing_files.append(file)
    print("Missing files:", missing_files)
    return missing_files

# Function to collect and log missing files
def collect_and_log_missing_files(root_folder):
    missing_files_global = []
    os.chdir(root_folder)
    # Traverse the directory structure
    for subdir in os.listdir(root_folder):
        print("subdirectory is :", subdir)
        missing_files = find_missing_files(subdir)
        print(missing_files)
        missing_files_global.extend(missing_files)
        print(missing_files)

    # Log all missing files in one consolidated text file
    output_log_file = os.path.join(root_folder, "missing_files_output.txt")
    with open(output_log_file, "w", encoding="utf-8") as log_file:
        for missing_file in missing_files_global:
            log_file.write(f"{missing_file}\n")

    print(f"Missing files have been logged in: {output_log_file}")

# Process each Excel file
print("Starting process_excel_files...")
folder_list = process_excel_files(directory)
print("process_excel_files completed.")
print(folder_list)
# Cleanup function
print("Starting cleanup function...")
for i in range(len(folder_list)):
    remove_zero_byte_files_and_empty_dirs(folder_list[i])
    
    print("one loop cleanup")
print("remove_zero_byte_files_and_empty_dirs completed.")
print(folder_list)
# Double-check and log missing files function
print("Starting collect_and_log_missing_files...")
for i in range(len(folder_list)):
    collect_and_log_missing_files(folder_list[i])
    print("one loop double check")
print("collect_and_log_missing_files completed.")

print("Python Program Terminated.")

